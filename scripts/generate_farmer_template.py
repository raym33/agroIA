from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_PATH = BASE_DIR / "templates" / "farmer_data_template_mvp.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="2F6F47")
SECTION_FILL = PatternFill("solid", fgColor="DDEBDC")
INFO_FILL = PatternFill("solid", fgColor="F8F3E7")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(color="FFFFFF", bold=True, size=16)
SUBTITLE_FONT = Font(color="2D2B26", bold=True, size=12)
BODY_FONT = Font(color="2D2B26", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D6D0C3"),
    right=Side(style="thin", color="D6D0C3"),
    top=Side(style="thin", color="D6D0C3"),
    bottom=Side(style="thin", color="D6D0C3"),
)


def style_headers(worksheet, row: int, columns: list[str]) -> None:
    for col_index, column_name in enumerate(columns, start=1):
        cell = worksheet.cell(row=row, column=col_index, value=column_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def style_body(worksheet, start_row: int, end_row: int, end_col: int) -> None:
    for row in worksheet.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=end_col):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top")
            if cell.row >= start_row and cell.column <= end_col and cell.value is None:
                cell.fill = INFO_FILL


def set_widths(worksheet, widths: dict[str, int]) -> None:
    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width


def make_table(worksheet, table_name: str, ref: str) -> None:
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def add_validation(worksheet, cell_range: str, formula: str) -> None:
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.prompt = "Select a value from the list."
    validation.error = "This value is not allowed for the field."
    worksheet.add_data_validation(validation)
    validation.add(cell_range)


def build_lists_sheet(workbook: Workbook) -> None:
    worksheet = workbook.create_sheet("Lists")
    worksheet.sheet_state = "hidden"

    worksheet["A1"] = "crop_type"
    worksheet["A2"] = "vegetables"
    worksheet["A3"] = "citrus"
    worksheet["A4"] = "orchard"

    worksheet["B1"] = "irrigation_type"
    worksheet["B2"] = "drip"
    worksheet["B3"] = "sprinkler"
    worksheet["B4"] = "flood"
    worksheet["B5"] = "other"

    worksheet["C1"] = "pest_observed"
    worksheet["C2"] = "yes"
    worksheet["C3"] = "no"

    worksheet["D1"] = "pest_risk_level"
    worksheet["D2"] = "low"
    worksheet["D3"] = "medium"
    worksheet["D4"] = "high"

    worksheet["E1"] = "treatment_type"
    worksheet["E2"] = "plant_protection"
    worksheet["E3"] = "fertilization"
    worksheet["E4"] = "irrigation"
    worksheet["E5"] = "other"

    worksheet["F1"] = "photo_type"
    worksheet["F2"] = "whole_plot"
    worksheet["F3"] = "leaf"
    worksheet["F4"] = "fruit"
    worksheet["F5"] = "harvest"
    worksheet["F6"] = "drone"


def build_instructions_sheet(workbook: Workbook) -> None:
    worksheet = workbook.active
    worksheet.title = "Instructions"
    worksheet.sheet_view.showGridLines = False
    worksheet.merge_cells("A1:H2")
    worksheet["A1"] = "Minimum farmer template - AgroIA local MVP"
    worksheet["A1"].fill = HEADER_FILL
    worksheet["A1"].font = TITLE_FONT
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

    content = [
        (
            "What to fill first",
            [
                "1. Plots sheet: one row per plot or farm block.",
                "2. FieldLog sheet: one row per day and plot.",
                "3. Treatments sheet: only when an action is performed.",
                "4. Photos sheet: file name and image context.",
            ],
        ),
        (
            "Key columns to start",
            [
                "Date, plot, crop, applied water, production, pest presence, and notes.",
                "If ET0, NDVI, or soil moisture do not exist yet, the project can still start and fill them later.",
            ],
        ),
        (
            "How to name photos",
            [
                "Use this format: PLOT-001_2026-05-02_leaf_01.jpg",
                "When possible, store photos in one folder per plot.",
            ],
        ),
        (
            "Purpose of this template",
            [
                "Understand which data already exists and which data is still missing.",
                "The farmer does not need perfect data from day one.",
            ],
        ),
    ]

    current_row = 4
    for section_title, lines in content:
        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        title_cell = worksheet.cell(row=current_row, column=1, value=section_title)
        title_cell.fill = SECTION_FILL
        title_cell.font = SUBTITLE_FONT
        title_cell.border = THIN_BORDER
        current_row += 1

        for line in lines:
            worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
            body_cell = worksheet.cell(row=current_row, column=1, value=line)
            body_cell.fill = INFO_FILL
            body_cell.font = BODY_FONT
            body_cell.alignment = Alignment(wrap_text=True, vertical="top")
            body_cell.border = THIN_BORDER
            current_row += 1

        current_row += 1

    set_widths(
        worksheet,
        {
            "A": 22,
            "B": 22,
            "C": 18,
            "D": 18,
            "E": 18,
            "F": 18,
            "G": 18,
            "H": 18,
        },
    )


def build_plot_sheet(workbook: Workbook) -> None:
    columns = [
        "plot_id",
        "plot_name",
        "crop_type",
        "area_ha",
        "municipality",
        "irrigation_type",
        "variety",
        "notes",
    ]
    worksheet = workbook.create_sheet("Plots")
    style_headers(worksheet, 1, columns)

    sample_rows = [
        ["PLOT-001", "North Farm", "vegetables", 1.8, "Torre Pacheco", "drip", "lettuce", "reference plot"],
        ["PLOT-002", "South Farm", "citrus", 3.2, "Murcia", "drip", "lemon", "sectorized irrigation"],
    ]
    for row_index, row_values in enumerate(sample_rows, start=2):
        for col_index, value in enumerate(row_values, start=1):
            worksheet.cell(row=row_index, column=col_index, value=value)

    style_body(worksheet, 2, 20, len(columns))
    add_validation(worksheet, "C2:C200", "=Lists!$A$2:$A$4")
    add_validation(worksheet, "F2:F200", "=Lists!$B$2:$B$5")
    worksheet.freeze_panes = "A2"
    set_widths(
        worksheet,
        {"A": 14, "B": 22, "C": 14, "D": 12, "E": 18, "F": 16, "G": 18, "H": 30},
    )
    make_table(worksheet, "plotsTable", "A1:H20")


def build_field_log_sheet(workbook: Workbook) -> None:
    columns = [
        "date",
        "plot_id",
        "temp_c",
        "humidity_pct",
        "et0_mm",
        "soil_moisture_pct",
        "rain_mm",
        "ndvi",
        "water_applied_l_m2",
        "pest_observed",
        "pest_risk_level",
        "production_kg",
        "notes",
    ]
    worksheet = workbook.create_sheet("FieldLog")
    style_headers(worksheet, 1, columns)

    sample_rows = [
        ["2026-05-02", "PLOT-001", 28.0, 47.0, 5.9, 21.0, 0.0, 0.72, 3.1, "no", "high", 1200, "outer leaves show mild stress"],
        ["2026-05-03", "PLOT-001", 27.4, 49.0, 5.4, 22.5, 0.0, 0.74, 2.8, "yes", "medium", 1230, "preventive review"],
    ]
    for row_index, row_values in enumerate(sample_rows, start=2):
        for col_index, value in enumerate(row_values, start=1):
            worksheet.cell(row=row_index, column=col_index, value=value)

    for cell in worksheet["A"][1:]:
        cell.number_format = "yyyy-mm-dd"
    style_body(worksheet, 2, 200, len(columns))
    add_validation(worksheet, "B2:B500", "=Plots!$A$2:$A$200")
    add_validation(worksheet, "J2:J500", "=Lists!$C$2:$C$3")
    add_validation(worksheet, "K2:K500", "=Lists!$D$2:$D$4")
    worksheet.freeze_panes = "A2"
    set_widths(
        worksheet,
        {
            "A": 14,
            "B": 14,
            "C": 10,
            "D": 14,
            "E": 10,
            "F": 18,
            "G": 10,
            "H": 10,
            "I": 18,
            "J": 14,
            "K": 16,
            "L": 14,
            "M": 32,
        },
    )
    make_table(worksheet, "fieldLogTable", "A1:M200")


def build_treatments_sheet(workbook: Workbook) -> None:
    columns = [
        "date",
        "plot_id",
        "type",
        "reason",
        "product",
        "area_pct",
        "cost_eur",
        "notes",
    ]
    worksheet = workbook.create_sheet("Treatments")
    style_headers(worksheet, 1, columns)

    sample_rows = [
        ["2026-05-02", "PLOT-001", "plant_protection", "aphids", "", 20, 35, "localized treatment"],
        ["2026-05-03", "PLOT-002", "irrigation", "schedule adjustment", "", 100, 0, "reduced by 30 minutes"],
    ]
    for row_index, row_values in enumerate(sample_rows, start=2):
        for col_index, value in enumerate(row_values, start=1):
            worksheet.cell(row=row_index, column=col_index, value=value)

    for cell in worksheet["A"][1:]:
        cell.number_format = "yyyy-mm-dd"
    style_body(worksheet, 2, 120, len(columns))
    add_validation(worksheet, "B2:B300", "=Plots!$A$2:$A$200")
    add_validation(worksheet, "C2:C300", "=Lists!$E$2:$E$5")
    worksheet.freeze_panes = "A2"
    set_widths(
        worksheet,
        {"A": 14, "B": 14, "C": 16, "D": 18, "E": 18, "F": 12, "G": 12, "H": 34},
    )
    make_table(worksheet, "treatmentsTable", "A1:H120")


def build_photos_sheet(workbook: Workbook) -> None:
    columns = [
        "date",
        "plot_id",
        "photo_name",
        "photo_type",
        "folder_path",
        "notes",
    ]
    worksheet = workbook.create_sheet("Photos")
    style_headers(worksheet, 1, columns)

    sample_rows = [
        ["2026-05-02", "PLOT-001", "PLOT-001_2026-05-02_leaf_01.jpg", "leaf", "PLOT-001/", "outer leaf with mild damage"],
        ["2026-05-02", "PLOT-001", "PLOT-001_2026-05-02_drone_01.jpg", "drone", "PLOT-001/", "full plot overview"],
    ]
    for row_index, row_values in enumerate(sample_rows, start=2):
        for col_index, value in enumerate(row_values, start=1):
            worksheet.cell(row=row_index, column=col_index, value=value)

    for cell in worksheet["A"][1:]:
        cell.number_format = "yyyy-mm-dd"
    style_body(worksheet, 2, 120, len(columns))
    add_validation(worksheet, "B2:B300", "=Plots!$A$2:$A$200")
    add_validation(worksheet, "D2:D300", "=Lists!$F$2:$F$6")
    worksheet.freeze_panes = "A2"
    set_widths(
        worksheet,
        {"A": 14, "B": 14, "C": 34, "D": 16, "E": 20, "F": 34},
    )
    make_table(worksheet, "photosTable", "A1:F120")


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()

    build_instructions_sheet(workbook)
    build_plot_sheet(workbook)
    build_field_log_sheet(workbook)
    build_treatments_sheet(workbook)
    build_photos_sheet(workbook)
    build_lists_sheet(workbook)

    workbook.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
